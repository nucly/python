import openpyxl as xl
from selenium import webdriver
from customer import Customer
import time


def load_customers(excel_file):

    cs_list = []
    customers = xl.load_workbook(path)
    ws = customers['Sheet1']

    for row in range(1, 3):
        for col in range(1, 7):

            firstname = ws.cell(row=row, column=1).value
            lastname = ws.cell(row=row, column=2).value
            dob = ws.cell(row=row, column=3).value
            address = ws.cell(row=row, column=4).value
            postcode = ws.cell(row=row, column=5).value
            email = ws.cell(row=row, column=6).value
            phone = ws.cell(row=row, column=7).value

            cs = Customer(firstname, lastname, dob,
                          address, postcode, email, phone)

        cs_list.append(cs)

    return cs_list


def send_to_html_form(customers):

    driver = webdriver.Chrome()

    for i in range(len(customers)):

        driver.get('http://127.0.0.1:5000/form')

        firstname = driver.find_element_by_id('firstname')
        lastname = driver.find_element_by_id('lastname')
        dob = driver.find_element_by_id('dob')
        address = driver.find_element_by_id('address')
        postcode = driver.find_element_by_id('postcode')
        email = driver.find_element_by_id('email')
        phone = driver.find_element_by_id('phone')

        firstname.send_keys(customers[i].first_name)
        lastname.send_keys(customers[i].last_name)
        dob.send_keys(customers[i].dob)
        address.send_keys(customers[i].address)
        postcode.send_keys(customers[i].postcode)
        email.send_keys(customers[i].email)
        phone.send_keys(customers[i].phone)
        driver.find_element_by_xpath("//input[@type='submit']").click();

    assert 'No result found' not in driver.page_source


if __name__ == '__main__':
    path = 'ExcelToHTMLForm/customers.xlsx'
    data = load_customers(path)
    print(data)
    send_to_html_form(data)


